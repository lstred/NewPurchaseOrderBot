"""
Reusable widget components shared across tabs.
"""

from __future__ import annotations

import csv
import os
import re
import sys
import tempfile
from datetime import datetime
from typing import Callable, Optional

from PyQt6.QtCore import Qt, QSize, QTimer, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QAction
from PyQt6.QtWidgets import (
    QFrame, QHBoxLayout, QLabel, QSizePolicy, QVBoxLayout, QWidget,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QLineEdit, QCheckBox, QScrollArea, QGroupBox,
    QMenu, QFileDialog, QMessageBox, QApplication,
)

import app.ui.theme as theme


_NUM_RE = re.compile(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+(\.\d+)?$")


def _coerce_excel_value(s):
    """Convert a stringified cell back to int/float when it looks numeric so
    Excel sorts/aggregates work natively. Leaves text and special markers
    (e.g. '—', '∞') untouched."""
    if not isinstance(s, str):
        return s
    raw = s.strip()
    if not raw or raw in ("—", "-", "∞", "N/A"):
        return raw
    if _NUM_RE.match(raw):
        try:
            cleaned = raw.replace(",", "")
            if "." in cleaned:
                return float(cleaned)
            return int(cleaned)
        except ValueError:
            return s
    return s


# ---------------------------------------------------------------------------
# Rule matching for DataTable threshold coloring
# ---------------------------------------------------------------------------

def _rule_matches(cell_val: str, op: str, threshold: str) -> bool:
    """Return True if cell_val satisfies the threshold rule."""
    # Strip non-numeric characters for numeric comparison
    clean = re.sub(r"[^0-9.\-]", "", cell_val)
    try:
        num = float(clean)
        thr = float(threshold)
        if op == ">":  return num > thr
        if op == ">=": return num >= thr
        if op == "<":  return num < thr
        if op == "<=": return num <= thr
        if op == "=":  return abs(num - thr) < 1e-9
        if op == "!=": return abs(num - thr) >= 1e-9
    except ValueError:
        pass
    # String fallback
    cv, tv = cell_val.strip().lower(), threshold.strip().lower()
    if op == "=":        return cv == tv
    if op == "!=":       return cv != tv
    if op == "contains": return tv in cv
    return False


def _contrasting_color(bg_hex: str) -> str:
    """Return #ffffff or #1a1a1a whichever contrasts better against bg_hex."""
    try:
        c = QColor(bg_hex)
        luminance = 0.299 * c.red() + 0.587 * c.green() + 0.114 * c.blue()
        return "#ffffff" if luminance < 128 else "#1a1a1a"
    except Exception:
        return "#ffffff"


class NumericTableWidgetItem(QTableWidgetItem):
    """QTableWidgetItem that sorts numerically when the cell text is a number.

    Handles formatted values: "101,845.8" → 101845.8, "2.76x" → 2.76,
    "100.0%" → 100.0, "∞" / "—" sort to the end.
    """

    def __lt__(self, other: "QTableWidgetItem") -> bool:
        try:
            return self._sort_key() < other._sort_key()  # type: ignore[attr-defined]
        except (AttributeError, TypeError):
            return self._sort_key() < self._key_from_text(other.text())

    def _sort_key(self):
        return self._key_from_text(self.text())

    @staticmethod
    def _key_from_text(text: str):
        s = text.strip()
        if s in ("—", "", "nan", "None"):
            return (2, 0.0)   # blanks/dashes sort last
        if s in ("∞", "inf", "INF"):
            return (1, 0.0)   # ∞ sorts second-to-last
        clean = re.sub(r"[^0-9.\-]", "", s)
        try:
            return (0, float(clean))
        except ValueError:
            return (2, s.lower())  # non-numeric: stable string sort


# ---------------------------------------------------------------------------
# Multi-condition rule evaluation helpers
# ---------------------------------------------------------------------------

def _eval_single_condition(row_data: list, column_names: list[str], cond: dict) -> bool:
    """Evaluate one {column, op, value} condition against a table row."""
    col = cond.get("column", "")
    try:
        idx = column_names.index(col)
    except ValueError:
        return False
    if idx >= len(row_data):
        return False
    cell_val = str(row_data[idx]) if row_data[idx] is not None else ""
    return _rule_matches(cell_val, cond.get("op", ">"), str(cond.get("value", "")))


def _eval_rule(row_data: list, column_names: list[str], rule: dict) -> bool:
    """Return True if ALL conditions in the rule match the given row.

    Supports new format (``rule["conditions"]`` list) and legacy flat
    format (``rule["column"]`` / ``rule["op"]`` / ``rule["value"]``).
    """
    conditions = rule.get("conditions")
    if conditions:
        return all(_eval_single_condition(row_data, column_names, c) for c in conditions)
    # Legacy single-condition format
    return _eval_single_condition(row_data, column_names, rule)


# ---------------------------------------------------------------------------
# KPI card
# ---------------------------------------------------------------------------

class KpiCard(QFrame):
    def __init__(self, label: str, value: str = "—", color_key: str = "accent", parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setMinimumWidth(160)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(4)

        self._value_label = QLabel(value)
        self._value_label.setObjectName("kpi_value")
        self._value_label.setStyleSheet(f"color: {theme.get(color_key)}; font-size: 28px; font-weight: 700;")

        self._label = QLabel(label.upper())
        self._label.setObjectName("kpi_label")

        layout.addWidget(self._value_label)
        layout.addWidget(self._label)

    def set_value(self, val: str, color_key: str = "accent") -> None:
        self._value_label.setText(val)
        self._value_label.setStyleSheet(f"color: {theme.get(color_key)}; font-size: 28px; font-weight: 700;")

    def refresh_theme(self) -> None:
        self.setStyleSheet("")  # trigger re-apply via parent QSS


# ---------------------------------------------------------------------------
# Section heading
# ---------------------------------------------------------------------------

class SectionTitle(QLabel):
    def __init__(self, text: str, parent=None):
        super().__init__(text, parent)
        self.setObjectName("section_title")
        font = QFont()
        font.setPointSize(14)
        font.setWeight(QFont.Weight.Bold)
        self.setFont(font)


# ---------------------------------------------------------------------------
# Horizontal rule separator
# ---------------------------------------------------------------------------

class HSep(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFrameShape(QFrame.Shape.HLine)
        self.setStyleSheet(f"color: {theme.get('border')};")
        self.setFixedHeight(1)


# ---------------------------------------------------------------------------
# Sortable table with column visibility and threshold-coloring rules
# ---------------------------------------------------------------------------

class DataTable(QTableWidget):
    def __init__(self, columns: list[str], parent=None, table_id: str | None = None):
        super().__init__(0, len(columns), parent)
        self._column_names: list[str] = list(columns)
        self._rules: list[dict] = []
        self._table_id: str | None = table_id
        self.setHorizontalHeaderLabels(columns)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setSortingEnabled(True)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.horizontalHeader().setStretchLastSection(True)
        self.horizontalHeader().setSectionsMovable(True)  # drag-to-reorder
        self.verticalHeader().setVisible(False)
        self.setShowGrid(True)

        # Persist column widths when user resizes — debounced 600 ms
        if table_id:
            self._width_timer = QTimer(self)
            self._width_timer.setSingleShot(True)
            self._width_timer.setInterval(600)
            self._width_timer.timeout.connect(self._save_column_widths)
            self.horizontalHeader().sectionResized.connect(
                lambda *_: self._width_timer.start()
            )

        # Right-click context menu: Open in Excel / Export / Copy
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)

    # ------------------------------------------------------------------
    # Export / Open in Excel
    # ------------------------------------------------------------------

    def _show_context_menu(self, pos) -> None:
        menu = QMenu(self)
        act_open = QAction("📊  Open in Excel", self)
        act_open.triggered.connect(self._open_in_excel)
        act_xlsx = QAction("💾  Export to Excel…", self)
        act_xlsx.triggered.connect(self._export_excel)
        act_csv = QAction("📄  Export to CSV…", self)
        act_csv.triggered.connect(self._export_csv)
        act_copy = QAction("📋  Copy selection", self)
        act_copy.setShortcut("Ctrl+C")
        act_copy.triggered.connect(self._copy_selection)
        menu.addAction(act_open)
        menu.addSeparator()
        menu.addAction(act_xlsx)
        menu.addAction(act_csv)
        menu.addSeparator()
        menu.addAction(act_copy)
        if self.rowCount() == 0:
            for a in (act_open, act_xlsx, act_csv):
                a.setEnabled(False)
        menu.exec(self.viewport().mapToGlobal(pos))

    def _visible_columns(self) -> list[int]:
        return [i for i in range(self.columnCount()) if not self.isColumnHidden(i)]

    def _table_snapshot(self) -> tuple[list[str], list[list[str]]]:
        """Capture currently visible columns + sorted rows as plain strings."""
        cols_idx = self._visible_columns()
        headers = [self._column_names[i] for i in cols_idx]
        rows: list[list[str]] = []
        for r in range(self.rowCount()):
            if self.isRowHidden(r):
                continue
            row: list[str] = []
            for i in cols_idx:
                it = self.item(r, i)
                row.append(it.text() if it is not None else "")
            rows.append(row)
        return headers, rows

    def _default_export_name(self, ext: str) -> str:
        stem = (self._table_id or "table").replace("/", "_")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{stem}_{ts}.{ext}"

    def _write_xlsx(self, path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        headers, rows = self._table_snapshot()
        wb = Workbook()
        ws = wb.active
        ws.title = (self._table_id or "Data")[:31]
        ws.append(headers)
        head_fill = PatternFill("solid", fgColor="1F2937")
        head_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for r in rows:
            ws.append([_coerce_excel_value(v) for v in r])
        ws.freeze_panes = "A2"
        # Auto-size columns (cap at 60 to keep widths sane)
        for col_idx, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for r in rows:
                if col_idx - 1 < len(r):
                    cell_len = len(str(r[col_idx - 1]))
                    if cell_len > max_len:
                        max_len = cell_len
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)
        wb.save(path)

    def _write_csv(self, path: str) -> None:
        headers, rows = self._table_snapshot()
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)

    def _open_in_excel(self) -> None:
        if self.rowCount() == 0:
            return
        try:
            tmp_dir = tempfile.gettempdir()
            stem = (self._table_id or "table").replace("/", "_")
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(tmp_dir, f"{stem}_{ts}.xlsx")
            try:
                self._write_xlsx(path)
            except ImportError:
                # Fallback to CSV if openpyxl missing
                path = os.path.join(tmp_dir, f"{stem}_{ts}.csv")
                self._write_csv(path)
            if sys.platform.startswith("win"):
                os.startfile(path)  # noqa: S606 — user-initiated
            elif sys.platform == "darwin":
                os.system(f'open "{path}"')  # noqa: S605
            else:
                os.system(f'xdg-open "{path}"')  # noqa: S605
        except Exception as e:  # noqa: BLE001
            QMessageBox.warning(self, "Open in Excel failed", f"{type(e).__name__}: {e}")

    def _export_excel(self) -> None:
        if self.rowCount() == 0:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", self._default_export_name("xlsx"),
            "Excel Workbook (*.xlsx)",
        )
        if not path:
            return
        try:
            self._write_xlsx(path)
        except ImportError:
            QMessageBox.warning(
                self, "openpyxl not installed",
                "The 'openpyxl' package is required for Excel export.\n\n"
                "Install it with:  pip install openpyxl\n\n"
                "Falling back to CSV — try 'Export to CSV…' instead.",
            )
        except Exception as e:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", f"{type(e).__name__}: {e}")

    def _export_csv(self) -> None:
        if self.rowCount() == 0:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export to CSV", self._default_export_name("csv"),
            "CSV file (*.csv)",
        )
        if not path:
            return
        try:
            self._write_csv(path)
        except Exception as e:  # noqa: BLE001
            QMessageBox.warning(self, "Export failed", f"{type(e).__name__}: {e}")

    def _copy_selection(self) -> None:
        ranges = self.selectedRanges()
        if not ranges:
            return
        rng = ranges[0]
        lines = []
        for r in range(rng.topRow(), rng.bottomRow() + 1):
            cells = []
            for c in range(rng.leftColumn(), rng.rightColumn() + 1):
                if self.isColumnHidden(c) or self.isRowHidden(r):
                    continue
                it = self.item(r, c)
                cells.append(it.text() if it is not None else "")
            lines.append("\t".join(cells))
        QApplication.clipboard().setText("\n".join(lines))

    def keyPressEvent(self, event) -> None:  # noqa: N802
        from PyQt6.QtGui import QKeySequence
        if event.matches(QKeySequence.StandardKey.Copy):
            self._copy_selection()
            return
        super().keyPressEvent(event)

    # ------------------------------------------------------------------
    # Dynamic columns (used by AI tab where columns vary per query)
    # ------------------------------------------------------------------

    def set_columns(self, columns: list[str]) -> None:
        """Replace the column set (clears any existing data)."""
        self.setSortingEnabled(False)
        self.clear()
        self.setRowCount(0)
        self._column_names = list(columns)
        self.setColumnCount(len(columns))
        self.setHorizontalHeaderLabels(columns)
        self.setSortingEnabled(True)

    # ------------------------------------------------------------------
    # Column visibility
    # ------------------------------------------------------------------

    def set_column_visible(self, col_name: str, visible: bool) -> None:
        """Show or hide a column by its header label."""
        try:
            idx = self._column_names.index(col_name)
            self.setColumnHidden(idx, not visible)
        except ValueError:
            pass

    def restore_column_widths(self) -> None:
        """Restore saved column widths (call after applying column visibility prefs)."""
        if not self._table_id:
            return
        from app.data.store import get_column_widths
        widths = get_column_widths(self._table_id)
        for col_name, width in widths.items():
            try:
                idx = self._column_names.index(col_name)
                self.setColumnWidth(idx, max(width, 20))
            except ValueError:
                pass

    def _save_column_widths(self) -> None:
        """Persist current column widths to disk."""
        if not self._table_id:
            return
        from app.data.store import set_column_widths
        widths = {
            col: self.columnWidth(i)
            for i, col in enumerate(self._column_names)
            if not self.isColumnHidden(i)
        }
        set_column_widths(self._table_id, widths)

    # ------------------------------------------------------------------
    # Threshold coloring rules
    # ------------------------------------------------------------------

    def set_rules(self, rules: list[dict]) -> None:
        """Set the coloring rules applied on every populate() call."""
        self._rules = list(rules)

    # ------------------------------------------------------------------
    # Data population with rule-based coloring
    # ------------------------------------------------------------------

    def populate(self, rows: list[list]) -> None:
        self.setUpdatesEnabled(False)
        self.setSortingEnabled(False)
        self.setRowCount(len(rows))
        for r, row_data in enumerate(rows):
            row_bg: str | None = None
            row_fg: str | None = None
            cell_overrides: dict[int, tuple[str | None, str | None]] = {}

            for rule in self._rules:
                if not _eval_rule(row_data, self._column_names, rule):
                    continue

                bg = rule.get("bg_color") or None
                fg = rule.get("fg_color") or None
                # Auto-contrast: if bg set but fg empty, pick readable text color
                if bg and not fg:
                    fg = _contrasting_color(bg)

                if rule.get("target") == "row":
                    row_bg, row_fg = bg, fg
                else:
                    apply_col = rule.get("apply_column", "")
                    if apply_col:
                        try:
                            apply_idx = self._column_names.index(apply_col)
                        except ValueError:
                            apply_idx = -1
                    else:
                        # Default: first condition's column (or legacy "column" key)
                        conditions = rule.get("conditions")
                        ref_col = (
                            conditions[0].get("column", "") if conditions
                            else rule.get("column", "")
                        )
                        try:
                            apply_idx = self._column_names.index(ref_col)
                        except ValueError:
                            apply_idx = -1
                    if apply_idx >= 0:
                        cell_overrides[apply_idx] = (bg, fg)

            for c, val in enumerate(row_data):
                item = NumericTableWidgetItem(str(val) if val is not None else "")
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

                if c in cell_overrides:
                    bg, fg = cell_overrides[c]
                else:
                    bg, fg = row_bg, row_fg

                if bg:
                    item.setBackground(QColor(bg))
                if fg:
                    item.setForeground(QColor(fg))

                self.setItem(r, c, item)
        self.setSortingEnabled(True)
        self.setUpdatesEnabled(True)


# ---------------------------------------------------------------------------
# Filter sidebar
# ---------------------------------------------------------------------------

class _CheckList(QWidget):
    """Scrollable group of checkboxes — replaces QListWidget multi-select."""

    changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._checks: dict[str, QCheckBox] = {}  # value → checkbox

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setMaximumHeight(115)
        self._scroll.setMinimumHeight(30)
        self._scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self._scroll.setFrameShape(QFrame.Shape.NoFrame)

        self._container = QWidget()
        self._inner = QVBoxLayout(self._container)
        self._inner.setContentsMargins(2, 2, 2, 2)
        self._inner.setSpacing(2)
        self._scroll.setWidget(self._container)
        lay.addWidget(self._scroll)

    def populate(self, items: list[tuple[str, str]]) -> None:
        """items: list of (display_label, filter_value)."""
        # Remove old widgets
        while self._inner.count():
            w = self._inner.takeAt(0).widget()
            if w:
                w.deleteLater()
        self._checks.clear()

        for label, value in items:
            cb = QCheckBox(label)
            cb.setProperty("_fv", value)
            cb.stateChanged.connect(self.changed)
            self._inner.addWidget(cb)
            self._checks[value] = cb

    def get_selected(self) -> list[str]:
        return [v for v, cb in self._checks.items() if cb.isChecked() and cb.isVisible()]

    def clear_all(self) -> None:
        for cb in self._checks.values():
            cb.blockSignals(True)
            cb.setChecked(False)
            cb.blockSignals(False)
        self.changed.emit()

    def has_selection(self) -> bool:
        return any(cb.isChecked() for cb in self._checks.values())

    def set_valid(self, valid_vals: set) -> None:
        """Hide and uncheck items not in valid_vals; show items that are valid."""
        for val, cb in self._checks.items():
            is_valid = val in valid_vals
            if not is_valid and cb.isChecked():
                cb.setChecked(False)
            cb.setVisible(is_valid)

    def show_all(self) -> None:
        """Make all items visible again (used on filter reset)."""
        for cb in self._checks.values():
            cb.setVisible(True)


class FilterSidebar(QFrame):
    filters_changed = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("sidebar")
        self.setFixedWidth(215)

        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        container = QWidget()
        self._layout = QVBoxLayout(container)
        self._layout.setContentsMargins(10, 12, 10, 12)
        self._layout.setSpacing(8)
        self._layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        # Debounce timer — only emit after 250 ms of inactivity
        self._timer = QTimer()
        self._timer.setSingleShot(True)
        self._timer.setInterval(250)
        self._timer.timeout.connect(self._emit)

        # Stored filter_values DataFrame for dynamic cascade updates
        self._full_fv = None

        # Title
        title = SectionTitle("Filters")
        self._layout.addWidget(title)

        # Search box
        search_gb = self._make_group_box("Search")
        self._sku_input = QLineEdit()
        self._sku_input.setPlaceholderText("SKU or description…")
        self._sku_input.textChanged.connect(self._schedule_emit)
        search_gb.layout().addWidget(self._sku_input)

        # Checkbox filter groups
        self._cc_list, _ = self._make_check_group("Cost Center")
        self._sup_list, _ = self._make_check_group("Supplier")
        self._pc_list, _ = self._make_check_group("Price Class")
        self._pl_list, _ = self._make_check_group("Product Line")

        # SKU Rating (fixed 4 options — horizontal checkboxes)
        rating_gb = self._make_group_box("SKU Rating")
        self._rating_checks: dict[str, QCheckBox] = {}
        row_lay = QHBoxLayout()
        row_lay.setSpacing(6)
        for r in ("A", "B", "C", "D"):
            cb = QCheckBox(r)
            cb.setChecked(True)
            cb.stateChanged.connect(self._schedule_emit)
            row_lay.addWidget(cb)
            self._rating_checks[r] = cb
        row_lay.addStretch()
        rating_gb.layout().addLayout(row_lay)

        # Reset all button
        btn_reset = QPushButton("Reset Filters")
        btn_reset.setObjectName("flat")
        btn_reset.clicked.connect(self._reset)
        self._layout.addWidget(btn_reset)
        self._layout.addStretch()

        scroll.setWidget(container)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

    # ------------------------------------------------------------------

    def _make_group_box(self, title: str) -> QGroupBox:
        gb = QGroupBox(title)
        vl = QVBoxLayout(gb)
        vl.setContentsMargins(6, 4, 6, 6)
        vl.setSpacing(4)
        self._layout.addWidget(gb)
        return gb

    def _make_check_group(self, title: str) -> tuple:
        """Group box with a Clear link + _CheckList. Returns (checklist, groupbox)."""
        gb = QGroupBox(title)
        vl = QVBoxLayout(gb)
        vl.setContentsMargins(6, 2, 6, 6)
        vl.setSpacing(2)

        hdr = QHBoxLayout()
        hdr.addStretch()
        clear_btn = QPushButton("Clear")
        clear_btn.setObjectName("link_btn")
        clear_btn.setFixedHeight(16)
        clear_btn.setStyleSheet(
            f"color: {theme.get('accent')}; border: none; background: transparent; "
            f"font-size: 11px; padding: 0;"
        )
        hdr.addWidget(clear_btn)
        vl.addLayout(hdr)

        cl = _CheckList()
        cl.changed.connect(self._on_filter_changed)
        clear_btn.clicked.connect(cl.clear_all)
        vl.addWidget(cl)

        self._layout.addWidget(gb)
        return cl, gb

    def populate(self, filter_values) -> None:
        import pandas as pd
        if filter_values is None or (isinstance(filter_values, pd.DataFrame) and filter_values.empty):
            return

        # Save for cascade filtering
        self._full_fv = filter_values

        # Cost centers — exclude those starting with '1' (internal use only)
        cc_vals = sorted({
            str(v) for v in filter_values["cost_center"].dropna().unique()
            if v and not str(v).startswith("1")
        })
        self._cc_list.populate([(v, v) for v in cc_vals])

        # Suppliers — no code-prefix exclusion
        sup_vals = sorted({
            str(v) for v in filter_values["supplier_number"].dropna().unique() if v
        })
        self._sup_list.populate([(v, v) for v in sup_vals])

        # Price classes — "CODE — Description"
        pc_df = filter_values[["price_class", "price_class_desc"]].drop_duplicates()
        pc_items: list[tuple[str, str]] = []
        for _, row in pc_df.dropna(subset=["price_class"]).iterrows():
            pc = str(row["price_class"]).strip()
            desc = str(row.get("price_class_desc", "")).strip()
            label = f"{pc} — {desc}" if desc and desc not in ("", "nan") else pc
            if pc:
                pc_items.append((label, pc))
        pc_items.sort()
        self._pc_list.populate(pc_items)

        # Product lines — "CODE — Description"
        pl_df = filter_values[["product_line", "product_line_desc"]].drop_duplicates()
        pl_items: list[tuple[str, str]] = []
        for _, row in pl_df.dropna(subset=["product_line"]).iterrows():
            pl = str(row["product_line"]).strip()
            desc = str(row.get("product_line_desc", "")).strip()
            label = f"{pl} — {desc}" if desc and desc not in ("", "nan") else pl
            if pl:
                pl_items.append((label, pl))
        pl_items.sort()
        self._pl_list.populate(pl_items)

    def get_filters(self) -> dict:
        return {
            "sku_search":   self._sku_input.text().strip(),
            "cost_centers": self._cc_list.get_selected() or None,
            "suppliers":    self._sup_list.get_selected() or None,
            "price_classes": self._pc_list.get_selected() or None,
            "product_lines": self._pl_list.get_selected() or None,
            "sku_ratings":  [r for r, cb in self._rating_checks.items() if cb.isChecked()],
        }

    def _on_filter_changed(self) -> None:
        """Called when any checklist checkbox changes: cascade then debounce."""
        self._update_dependent_filters()
        self._timer.start()

    def _schedule_emit(self) -> None:
        """Debounce for the search box (no cascade needed)."""
        self._timer.start()

    def _emit(self) -> None:
        self.filters_changed.emit(self.get_filters())

    def _update_dependent_filters(self) -> None:
        """Narrow each filter group to only options compatible with other active selections."""
        import pandas as pd

        fv = self._full_fv
        if fv is None or not isinstance(fv, pd.DataFrame) or fv.empty:
            return

        cc_sel  = set(self._cc_list.get_selected())
        sup_sel = set(self._sup_list.get_selected())
        pc_sel  = set(self._pc_list.get_selected())
        pl_sel  = set(self._pl_list.get_selected())

        # Block all checkbox signals to prevent re-entrancy during update
        all_cbs = [
            cb
            for cl in (self._cc_list, self._sup_list, self._pc_list, self._pl_list)
            for cb in cl._checks.values()
        ]
        for cb in all_cbs:
            cb.blockSignals(True)

        try:
            if not any((cc_sel, sup_sel, pc_sel, pl_sel)):
                # Nothing selected — show everything
                for cl in (self._cc_list, self._sup_list, self._pc_list, self._pl_list):
                    cl.show_all()
            else:
                cc_valid  = self._compute_valid(
                    fv, "cost_center",
                    ("supplier_number", sup_sel), ("price_class", pc_sel), ("product_line", pl_sel))
                sup_valid = self._compute_valid(
                    fv, "supplier_number",
                    ("cost_center", cc_sel), ("price_class", pc_sel), ("product_line", pl_sel))
                pc_valid  = self._compute_valid(
                    fv, "price_class",
                    ("cost_center", cc_sel), ("supplier_number", sup_sel), ("product_line", pl_sel))
                pl_valid  = self._compute_valid(
                    fv, "product_line",
                    ("cost_center", cc_sel), ("supplier_number", sup_sel), ("price_class", pc_sel))

                self._cc_list.set_valid(cc_valid)
                self._sup_list.set_valid(sup_valid)
                self._pc_list.set_valid(pc_valid)
                # Product line: if no non-empty values found (items have no PL assigned),
                # show all options rather than hiding everything.
                if pl_valid:
                    self._pl_list.set_valid(pl_valid)
                else:
                    self._pl_list.show_all()
        finally:
            for cb in all_cbs:
                cb.blockSignals(False)

    def _compute_valid(self, fv, dim: str, *constraints) -> set:
        """Return valid non-empty values for ``dim`` after applying all other active selections."""
        import pandas as pd

        mask = pd.Series(True, index=fv.index)
        for col, sel in constraints:
            if sel and col in fv.columns:
                mask &= fv[col].isin(sel)
        # Strip empty strings — items with no value for this dimension should not
        # cause the whole checklist to collapse to zero options.
        return {v for v in fv.loc[mask, dim].dropna().astype(str).str.strip() if v}

    def _reset(self) -> None:
        """Clear all selections and restore all filter items."""
        self._sku_input.blockSignals(True)
        self._sku_input.clear()
        self._sku_input.blockSignals(False)

        all_cbs = [
            cb
            for cl in (self._cc_list, self._sup_list, self._pc_list, self._pl_list)
            for cb in cl._checks.values()
        ]
        for cb in all_cbs:
            cb.blockSignals(True)
        try:
            for cl in (self._cc_list, self._sup_list, self._pc_list, self._pl_list):
                for cb in cl._checks.values():
                    cb.setChecked(False)
                    cb.setVisible(True)
        finally:
            for cb in all_cbs:
                cb.blockSignals(False)

        for cb in self._rating_checks.values():
            cb.blockSignals(True)
            cb.setChecked(True)
            cb.blockSignals(False)

        self._emit()


# ---------------------------------------------------------------------------
# Status badge
# ---------------------------------------------------------------------------

def make_badge(text: str, color_key: str = "accent") -> QLabel:
    lbl = QLabel(text)
    lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
    lbl.setStyleSheet(
        f"background:{theme.get(color_key)}; color:#fff; border-radius:4px;"
        "padding: 2px 8px; font-size: 11px; font-weight: 600;"
    )
    return lbl


# ---------------------------------------------------------------------------
# Plotly chart widget (renders HTML in QWebEngineView if available,
# falls back to a placeholder label)
# ---------------------------------------------------------------------------

def make_chart_widget(fig=None, parent=None) -> QWidget:
    try:
        from PyQt6.QtWebEngineWidgets import QWebEngineView
        view = QWebEngineView(parent)
        if fig is not None:
            html = fig.to_html(include_plotlyjs="cdn", full_html=True)
            view.setHtml(html)
        return view
    except ImportError:
        lbl = QLabel("Install PyQt6-WebEngine to display charts.", parent)
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setStyleSheet(f"color: {theme.get('text_muted')};")
        return lbl


def update_chart_widget(widget: QWidget, fig) -> None:
    try:
        from PyQt6.QtWebEngineWidgets import QWebEngineView
        if isinstance(widget, QWebEngineView):
            html = fig.to_html(include_plotlyjs="cdn", full_html=True)
            widget.setHtml(html)
    except ImportError:
        pass
